import unittest
from pathlib import Path
import sys
import types
import shutil

from openpyxl import load_workbook

# Make XLSX generator import independent from runtime-only settings dependency.
if "app.config" not in sys.modules:
    fake_config = types.ModuleType("app.config")
    fake_config.settings = types.SimpleNamespace(REPORTS_DIR=".")
    sys.modules["app.config"] = fake_config

from app.reports.xlsx_generator import XLSXGenerator


class SiteProXlsxLayoutTests(unittest.TestCase):
    def test_full_report_sheet_and_header_contract_is_stable(self):
        data = {
            "url": "https://site.test",
            "results": {
                "engine": "site_pro_adapter_v0",
                "mode": "full",
                "summary": {
                    "total_pages": 1,
                    "internal_pages": 1,
                    "issues_total": 1,
                    "critical_issues": 0,
                    "warning_issues": 1,
                    "info_issues": 0,
                    "score": 91.0,
                },
                "pages": [
                    {
                        "url": "https://site.test",
                        "final_url": "https://site.test",
                        "status_code": 200,
                        "response_time_ms": 220,
                        "html_size_bytes": 15320,
                        "dom_nodes_count": 340,
                        "redirect_count": 0,
                        "is_https": True,
                        "compression_enabled": True,
                        "cache_enabled": True,
                        "indexable": True,
                        "health_score": 91.0,
                        "title": "Home",
                        "meta_description": "Main page",
                        "canonical": "https://site.test",
                        "meta_robots": "index,follow",
                        "x_robots_tag": "",
                        "schema_count": 1,
                        "structured_data": 1,
                        "structured_errors_count": 0,
                        "structured_error_codes": [],
                        "structured_types": ["Organization"],
                        "hreflang_count": 0,
                        "hreflang_issues": [],
                        "mobile_friendly_hint": True,
                        "word_count": 330,
                        "unique_word_count": 180,
                        "lexical_diversity": 0.545,
                        "readability_score": 84.2,
                        "toxicity_score": 0.0,
                        "filler_ratio": 0.03,
                        "h1_count": 1,
                        "images_count": 8,
                        "images_without_alt": 1,
                        "external_nofollow_links": 1,
                        "external_follow_links": 2,
                        "outgoing_internal_links": 6,
                        "incoming_internal_links": 3,
                        "outgoing_external_links": 3,
                        "orphan_page": False,
                        "topic_hub": True,
                        "pagerank": 100.0,
                        "topic_label": "home",
                        "top_terms": ["home", "services", "seo"],
                        "top_keywords": ["home"],
                        "duplicate_title_count": 1,
                        "duplicate_description_count": 1,
                        "weak_anchor_ratio": 0.08,
                        "link_quality_score": 96.5,
                        "ai_markers_count": 0,
                        "ai_markers_list": [],
                        "ai_marker_sample": "",
                        "recommendation": "Maintain page quality and monitor regressions.",
                        "issues": [{"severity": "warning", "code": "thin_content", "title": "Thin", "details": ""}],
                        "path_depth": 0,
                        "url_params_count": 0,
                        "crawl_budget_risk": "low",
                        "near_duplicate_count": 0,
                        "perf_light_score": 88.0,
                        "eeat_score": 70.0,
                        "trust_score": 72.0,
                        "eeat_components": {
                            "expertise": 70,
                            "authoritativeness": 68,
                            "trustworthiness": 74,
                            "experience": 69,
                        },
                        "has_author_info": True,
                        "has_contact_info": True,
                        "has_legal_docs": True,
                        "has_reviews": False,
                        "trust_badges": False,
                        "entity_consistency_score": 80.0,
                        "page_type": "home",
                        "ai_risk_score": 10.0,
                        "ai_risk_level": "low",
                        "ai_false_positive_guard": True,
                        "broken_internal_targets": [],
                        "semantic_links": [
                            {"target_url": "https://site.test/about", "topic": "about", "reason": "related"}
                        ],
                        "editorial_policy_present": True,
                        "sources_cited": True,
                    }
                ],
                "issues": [
                    {"severity": "warning", "url": "https://site.test", "code": "thin_content", "title": "Thin", "details": ""}
                ],
                "pipeline": {
                    "tf_idf": [{"url": "https://site.test", "top_terms": ["home", "services", "seo"]}],
                    "semantic_linking_map": [
                        {
                            "source_url": "https://site.test",
                            "target_url": "https://site.test/about",
                            "topic": "about",
                            "reason": "related",
                        }
                    ],
                    "duplicates": {"title_groups": [], "description_groups": []},
                    "metrics": {
                        "avg_response_time_ms": 220.0,
                        "avg_readability_score": 84.2,
                        "avg_link_quality_score": 96.5,
                        "avg_perf_light_score": 88.0,
                        "orphan_pages": 0,
                        "topic_hubs": 1,
                        "pages_without_alt": 1,
                        "non_https_pages": 0,
                        "crawl_budget_high_risk": 0,
                        "crawl_budget_medium_risk": 0,
                    },
                    "pagerank": [{"url": "https://site.test", "score": 100.0}],
                    "topic_clusters": [{"topic": "home", "urls": ["https://site.test"], "count": 1}],
                },
            },
        }

        expected_headers = {
            "1_Executive": ["Отчет Site Audit Pro"],
            "2_OnPage+Structured": ["URL", "Title", "Длина title", "Ширина title, px", "Meta description", "Длина description", "Ширина description, px", "Риск обрезки в SERP", "Кол-во H1", "Текст H1", "Canonical URL", "Статус canonical", "Meta robots", "X-Robots", "Schema count", "JSON-LD", "Microdata", "RDFa", "Типы structured", "Кол-во hreflang", "Хлебные крошки", "Mobile hint", "Charset", "Viewport", "Дубликаты meta robots", "Кол-во title-тегов", "Кол-во description-тегов", "Дубли title", "Дубли description", "Self-match canonical", "OnPage-скор", "OnPage-дельта до цели", "OnPage-решение", "Severity"],
            "3_Technical": ["URL", "Итоговый URL", "Status", "Строка статуса", "Response ms", "Размер, KB", "HTML байт", "DOM-узлы", "Редиректы", "HTTPS", "Compression", "Compression algo", "Cache enabled", "Cache-Control", "Last-Modified", "Freshness days", "JS assets", "CSS assets", "Render-blocking JS", "Preload hints", "Perf light score", "Path depth", "URL params", "Crawl budget risk", "Security headers score", "CSP", "HSTS", "X-Frame-Options", "Referrer-Policy", "Permissions-Policy", "Mixed content refs", "Оценка качества HTML", "Кол-во устаревших тегов", "Причина неиндексируемости", "TTFB, мс", "Соотношение HTML/JS", "Риск цепочки редиректов", "Транспортный риск", "Уровень транспорта", "Технический скор", "Техническая дельта до цели", "Техническое решение", "Severity"],
            "4_Content+AI": ["URL", "Word count", "Уникальных слов", "Уникальность, %", "Лексическое разнообразие", "Скор читабельности", "Ср. длина предложения", "Ср. длина слова", "Сложные слова, %", "Скор keyword stuffing", "Плотность контента, %", "Boilerplate, %", "Toxicity-скор", "Доля «воды»", "Фразы-наполнители", "Кол-во AI-маркеров", "Список AI-маркеров", "Пример AI-маркеров", "Плотность AI /1k", "AI-риск", "Уровень AI-риска", "Тип страницы", "Скрытый контент", "Скрытые узлы", "Символов скрытого текста", "Клоакинг", "Соотношение контент/шаблон", "Кол-во абзацев", "Ср. длина абзаца", "Критичность скрытого", "Кол-во CTA", "Качество CTA", "Микс типов CTA", "Кол-во списков", "Кол-во таблиц", "Близкие дубли", "URL близких дублей", "Контент-скор", "Контент-дельта до цели", "Контент-решение", "Severity"],
            "5_LinkGraph": ["URL", "Входящие внутренние", "Исходящие внутренние", "Исходящие внешние", "Страница-сирота", "Тематический хаб", "Глубина клика", "PageRank", "Доля слабых анкоров", "Качество анкоров", "Авторитетность страницы", "Качество ссылок", "Всего follow-ссылок", "External nofollow", "Кол-во семантических ссылок", "Возможности внутренней перелинковки", "Битые внутренние цели", "Алерт переиспользования анкоров", "Link-скор", "Link-дельта до цели", "Решение по перелинковке", "Severity"],
            "6_Images+External": ["URL", "Всего изображений", "Без alt", "Без width/height", "Без lazy-load", "Всего image-проблем", "Современные форматы", "Дубли src", "Внешние изображения", "Generic ALT", "Декоративные с ALT", "Всего внешних", "Внешние follow", "Внешние nofollow", "Доля follow, %", "Макс. размер изображения, KB", "Без srcset", "Домены внешних изображений", "Релевантность ALT", "Медиа-скор", "Медиа-дельта до цели", "Решение по изображениям+внешним", "Severity"],
            "7_HierarchyErrors": ["URL", "Статус иерархии", "Проблемы иерархии", "Всего заголовков", "H1 count (Hierarchy)", "Outline заголовков", "Code", "Заголовок проблемы", "Детали проблемы", "H2 до H1", "Скор глубины outline", "Дубли текстов заголовков", "TOC-ready", "Hierarchy-скор", "Hierarchy-дельта до цели", "Решение по иерархии", "Severity"],
            "8_Keywords": ["URL", "Тема", "Top terms (TF-IDF)", "Топ ключей", "TF-IDF #1", "TF-IDF #2", "TF-IDF #3", "Профиль плотности ключей", "TF-IDF термины", "Энтропия ключей", "Доля топ-ключа, %", "SPAM-алерт", "Вода, %", "BM25-подобная релевантность", "Точное в Title", "Точное в H1", "Точное в URL", "Уверенность интента", "Уровень интента", "Keyword-скор", "Keyword-дельта до цели", "Решение по ключам", "Severity"],
            "8b_Keywords_Summary": ["N-gram", "Ключ", "Общая частота", "Страниц с термином", "Страниц %", "Доля токена %", "Средний TF-IDF", "Пиковая плотность %", "SPAM-сигнал", "Вода/шум", "Межстраничный повтор", "Риск-скор", "Брендовый термин", "Интент термина", "Сводная заметка"],
            "8c_Keywords_Insights": ["Метод", "Область", "Интент", "Термин/паттерн", "Метрика", "Критичность", "Действие", "Примеры", "Приоритет"],
            "9_Indexability": ["URL", "Статус", "Индексируемо", "Noindex", "Заблокировано robots", "Причина индексируемости", "Canonical URL", "Статус canonical", "Meta robots", "X-Robots-Tag", "Тип конфликта", "В sitemap", "Риск обнаружения", "Скор индексируемости", "Уровень индексируемости", "Дельта индексируемости до цели", "Решение по индексируемости", "Критичность"],
            "10_StructuredData": ["URL", "Всего structured", "JSON-LD", "Microdata", "RDFa", "Типы structured", "Hreflang", "Хлебные крошки", "FAQ schema", "Product schema", "Article schema", "Конфликт schema", "Подходит для rich result", "Критичные schema-ошибки", "Кол-во типовых ошибок", "Коды типовых ошибок", "Покрытие structured, %", "Structured-дельта до цели", "Решение по structured", "Критичность"],
            "11_Trust_EEAT": ["URL", "Trust-скор", "EEAT-скор", "Экспертиза", "Авторитет", "Надежность", "Опыт", "Инфо об авторе", "Контакты", "Юридическая инфо", "Отзывы", "Бейджи", "Редакционная политика", "Указанные источники", "EEAT-матрица", "Trust-gap", "Кол-во trust-доказательств", "Trust-дельта до цели", "Решение Trust+EEAT", "Критичность"],
            "12_Topics_Semantics": ["URL", "Тема", "Хаб", "Входящие ссылки", "Исходящие внутренние ссылки", "Кол-во семантических ссылок", "Рекомендуемые ссылки", "Детали семантических ссылок", "Топ терминов", "Топ ключей", "Скор тематической глубины", "Тематическая дельта до цели", "Полнота кластера", "Сиротский узел темы", "Перегруз хаба", "Консистентность сущностей", "Решение по темам", "Критичность"],
            "13_AI_Markers": ["URL", "AI-маркеры", "Список AI-маркеров", "Пример маркеров", "Плотность AI /1k", "AI risk-скор", "Уровень AI-риска", "Защита от false-positive", "Тип страницы", "Toxicity-скор", "Доля «воды»", "Стилевые маркеры", "Disclaimer-маркеры", "Transition-маркеры", "Hedging-маркеры", "Уверенность false-positive", "AI-риск выше порога", "Подсказка по humanization", "Критичность"],
            "CrawlBudget": ["URL", "Глубина пути", "Параметры URL", "Риск crawl budget", "Редиректы", "Статус", "Индексируемо", "Входящие ссылки", "Исходящие внутренние", "Близкие дубли", "Дубли обхода сверх цели", "Группа параметров", "Скор потерь обхода", "Риск глубокой индексируемости", "Решение по crawl budget", "Критичность"],
            "14_Issues_Raw": ["Критичность", "URL", "Код", "Категория", "Первая вкладка", "Dedupe hash", "Ответственный", "Заголовок", "Детали", "Затронуто", "Рекомендация"],
            "15_ActionPlan": ["Приоритет", "Код проблемы", "Макс. критичность", "Затронуто страниц", "Доля %", "Критично", "Предупреждение", "Инфо", "Импакт-скор", "Трудоемкость", "ETA", "Подсказка владельца", "Ожидаемый эффект", "Кластер первопричин", "Зависит от кодов", "Зависимость", "Потенциал batch-фикса", "ROI-скор", "Бакет спринта", "Репрезентативные URL", "Рекомендация"],
        }

        temp_dir = Path("tests") / ".tmp_site_pro_xlsx_contract"
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            generator = XLSXGenerator()
            generator.reports_dir = str(temp_dir)
            report_path = generator.generate_site_audit_pro_report("site-pro-layout-contract", data)
            wb = load_workbook(report_path)

            self.assertEqual(wb.sheetnames, list(expected_headers.keys()))
            for sheet_name, headers in expected_headers.items():
                ws = wb[sheet_name]
                actual_headers = [cell.value for cell in ws[1] if cell.value is not None]
                self.assertEqual(actual_headers, headers, msg=f"Unexpected header contract drift in sheet '{sheet_name}'")
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_quick_report_has_8_sheets_and_unique_parameter_headers(self):
        data = {
            "url": "https://site.test",
            "results": {
                "mode": "quick",
                "summary": {
                    "total_pages": 1,
                    "internal_pages": 1,
                    "issues_total": 1,
                    "critical_issues": 0,
                    "warning_issues": 1,
                    "info_issues": 0,
                    "score": 91.0,
                },
                "pages": [
                    {
                        "url": "https://site.test",
                        "final_url": "https://site.test",
                        "status_code": 200,
                        "response_time_ms": 220,
                        "html_size_bytes": 15320,
                        "dom_nodes_count": 340,
                        "redirect_count": 0,
                        "is_https": True,
                        "compression_enabled": True,
                        "cache_enabled": True,
                        "indexable": True,
                        "health_score": 91.0,
                        "title": "Home",
                        "meta_description": "Main page",
                        "canonical": "https://site.test",
                        "meta_robots": "index,follow",
                        "schema_count": 1,
                        "hreflang_count": 0,
                        "mobile_friendly_hint": True,
                        "word_count": 330,
                        "unique_word_count": 180,
                        "lexical_diversity": 0.545,
                        "readability_score": 84.2,
                        "toxicity_score": 0.0,
                        "filler_ratio": 0.03,
                        "h1_count": 1,
                        "images_count": 8,
                        "images_without_alt": 1,
                        "external_nofollow_links": 1,
                        "external_follow_links": 2,
                        "outgoing_internal_links": 6,
                        "incoming_internal_links": 3,
                        "outgoing_external_links": 3,
                        "orphan_page": False,
                        "topic_hub": True,
                        "pagerank": 100.0,
                        "topic_label": "home",
                        "top_terms": ["home", "services", "seo"],
                        "duplicate_title_count": 1,
                        "duplicate_description_count": 1,
                        "weak_anchor_ratio": 0.08,
                        "link_quality_score": 96.5,
                        "ai_markers_count": 0,
                        "recommendation": "Maintain page quality and monitor regressions.",
                        "issues": [{"severity": "warning", "code": "thin_content", "title": "Thin", "details": ""}],
                    }
                ],
                "issues": [
                    {"severity": "warning", "url": "https://site.test", "code": "thin_content", "title": "Thin", "details": ""}
                ],
                "pipeline": {
                    "tf_idf": [{"url": "https://site.test", "top_terms": ["home", "services", "seo"]}]
                },
            },
        }

        temp_dir = Path("tests") / ".tmp_site_pro_xlsx"
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            generator = XLSXGenerator()
            generator.reports_dir = str(temp_dir)
            report_path = generator.generate_site_audit_pro_report("site-pro-layout-test", data)

            wb = load_workbook(report_path)
            expected_sheets = [
                "1_Executive",
                "2_OnPage+Structured",
                "3_Technical",
                "4_Content+AI",
                "5_LinkGraph",
                "6_Images+External",
                "7_HierarchyErrors",
                "8_Keywords",
            ]
            self.assertEqual(wb.sheetnames[:8], expected_sheets)

            allow_repeats = {"URL", "Severity"}
            used_headers = {}
            for sheet_name in expected_sheets[1:]:
                ws = wb[sheet_name]
                headers = [cell.value for cell in ws[1] if cell.value]
                for header in headers:
                    if header in allow_repeats:
                        continue
                    self.assertNotIn(
                        header,
                        used_headers,
                        msg=f"Header '{header}' is duplicated in '{used_headers.get(header)}' and '{sheet_name}'",
                    )
                    used_headers[header] = sheet_name

            # Smoke-check that key metric families are represented exactly once.
            required_headers = {
                "Status",
                "Response ms",
                "Schema count",
                "Word count",
                "PageRank",
                "External nofollow",
                "Code",
                "Top terms (TF-IDF)",
            }
            self.assertTrue(required_headers.issubset(set(used_headers.keys())))

            self.assertTrue(Path(report_path).exists())
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_full_report_adds_deep_sheets(self):
        data = {
            "url": "https://site.test",
            "results": {
                "mode": "full",
                "summary": {"total_pages": 1, "issues_total": 0, "critical_issues": 0, "warning_issues": 0, "info_issues": 0},
                "pages": [{"url": "https://site.test", "topic_label": "home", "top_terms": ["home"], "issues": []}],
                "issues": [],
                "pipeline": {
                    "tf_idf": [{"url": "https://site.test", "top_terms": ["home"]}],
                    "semantic_linking_map": [
                        {"source_url": "https://site.test", "target_url": "https://site.test/a", "topic": "home", "reason": "test"}
                    ],
                    "duplicates": {"title_groups": [], "description_groups": []},
                },
            },
        }

        temp_dir = Path("tests") / ".tmp_site_pro_xlsx_full"
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            generator = XLSXGenerator()
            generator.reports_dir = str(temp_dir)
            report_path = generator.generate_site_audit_pro_report("site-pro-layout-full", data)
            wb = load_workbook(report_path)
            self.assertIn("9_Indexability", wb.sheetnames)
            self.assertIn("10_StructuredData", wb.sheetnames)
            self.assertIn("11_Trust_EEAT", wb.sheetnames)
            self.assertIn("12_Topics_Semantics", wb.sheetnames)
            self.assertIn("13_AI_Markers", wb.sheetnames)
            self.assertIn("CrawlBudget", wb.sheetnames)
            self.assertIn("14_Issues_Raw", wb.sheetnames)
            self.assertIn("15_ActionPlan", wb.sheetnames)

            self.assertNotIn("13_MainReport_Compat", wb.sheetnames)
            self.assertNotIn("29_AIMarkers_Compat", wb.sheetnames)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
